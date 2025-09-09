"""
Create Excel workbook comparing indicators and signals
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from main_daily_backtester import MainDailyBacktester

def create_fixed_comparison_workbook():
    """Create Excel workbook comparing Excel formulas vs Python calculations"""
    
    print("Creating FIXED comparison workbook...")
    
    # Get Python-generated data
    backtester = MainDailyBacktester()
    results = backtester.run_strategy_backtest('top_v2')
    python_data = results['data_with_signals']
    
    # Load raw data
    raw_data = pd.read_csv('top_v2_daily_data_2022.csv')
    raw_data['date'] = pd.to_datetime(raw_data['date'])
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    excel_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    python_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    signal_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Create main comparison sheet
    ws = wb.create_sheet("Excel vs Python")
    
    # Updated headers with proper signal logic description
    headers = [
        "Date", "ES_Close", "VIX_Close", "TRIN_Close", "CNN_Fear_Greed", "ES_RSI_2_Excel", "ES_EMA_9_Excel", "ES_EMA_15_Excel",
        "Excel_Base1_CNN<30", "Excel_Base2_RSI>50", "Excel_Cond1_VIX>20", "Excel_Cond2_RSI>60", "Excel_Cond3_TRIN<0.9", 
        "Excel_Cond_Sum", "Excel_Group_Signal", "Excel_Entry_Signal",
        "ES_RSI_2_Python", "ES_EMA_9_Python", "ES_EMA_15_Python", "Python_Entry_Signal",
        "Signal_Match"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Write data row by row
    for idx, (raw_idx, raw_row) in enumerate(raw_data.iterrows(), 2):
        # Find corresponding Python data
        py_row = python_data[python_data['date'] == raw_row['date']]
        if py_row.empty:
            continue
        py_row = py_row.iloc[0]
        
        # Raw data columns (A-F)
        ws.cell(row=idx, column=1, value=raw_row['date'])
        ws.cell(row=idx, column=2, value=raw_row['ES_close'] if pd.notna(raw_row['ES_close']) else "")
        ws.cell(row=idx, column=3, value=raw_row['VIX_close'] if pd.notna(raw_row['VIX_close']) else "")
        ws.cell(row=idx, column=4, value=raw_row['TRIN_close'] if pd.notna(raw_row['TRIN_close']) else "")
        ws.cell(row=idx, column=5, value=raw_row['CNN_FEAR_GREED'] if pd.notna(raw_row['CNN_FEAR_GREED']) else "")
        # Excel RSI - reference from RSI Debug sheet (will be calculated there properly)
        ws.cell(row=idx, column=6, value=f'=IF(\'RSI Debug\'!I{idx}="", "", \'RSI Debug\'!I{idx})')
        
        # Excel EMA calculations - we'll add these as formulas after all data is written
        ws.cell(row=idx, column=7, value="")  # ES_EMA_9_Excel
        ws.cell(row=idx, column=8, value="")  # ES_EMA_15_Excel
        
        # Excel formulas (I-O) - using DIRECT cell references, OUTPUT 1/0 like Python
        # Base Rule 1: CNN < 30 (output 1/0)
        ws.cell(row=idx, column=9, value=f'=IF(E{idx}="", 0, IF(E{idx}<30, 1, 0))')
        ws.cell(row=idx, column=9).fill = excel_fill
        
        # Base Rule 2: RSI > 50 (output 1/0)
        ws.cell(row=idx, column=10, value=f'=IF(F{idx}="", 0, IF(F{idx}>50, 1, 0))')
        ws.cell(row=idx, column=10).fill = excel_fill
        
        # Conditional Rule 1: VIX > 20 (output 1/0)
        ws.cell(row=idx, column=11, value=f'=IF(C{idx}="", 0, IF(C{idx}>20, 1, 0))')
        ws.cell(row=idx, column=11).fill = signal_fill
        
        # Conditional Rule 2: RSI > 60 (output 1/0)
        ws.cell(row=idx, column=12, value=f'=IF(F{idx}="", 0, IF(F{idx}>60, 1, 0))')
        ws.cell(row=idx, column=12).fill = signal_fill
        
        # Conditional Rule 3: TRIN < 0.9 (output 1/0)
        ws.cell(row=idx, column=13, value=f'=IF(D{idx}="", 0, IF(D{idx}<0.9, 1, 0))')
        ws.cell(row=idx, column=13).fill = signal_fill
        
        # Conditional Sum (sum of the 3 conditional rules)
        ws.cell(row=idx, column=14, value=f'=K{idx}+L{idx}+M{idx}')
        ws.cell(row=idx, column=14).fill = signal_fill
        
        # Conditional Group Signal: (CondSum >= 1) -> 1/0 (matching Python's group_signal logic)
        ws.cell(row=idx, column=15, value=f'=IF(N{idx}>=1, 1, 0)')
        ws.cell(row=idx, column=15).fill = signal_fill
        
        # Final Excel Entry Signal: Base1=1 AND Base2=1 AND GroupSignal=1 (output 1/0)
        ws.cell(row=idx, column=16, value=f'=IF(AND(I{idx}=1, J{idx}=1, O{idx}=1), 1, 0)')
        ws.cell(row=idx, column=16).fill = signal_fill
        
        # Python indicator values (Q-T)
        ws.cell(row=idx, column=17, value=py_row['ES_RSI_2'] if pd.notna(py_row['ES_RSI_2']) else "")
        ws.cell(row=idx, column=17).fill = python_fill
        
        ws.cell(row=idx, column=18, value=py_row['ES_EMA_9'] if pd.notna(py_row['ES_EMA_9']) else "") 
        ws.cell(row=idx, column=18).fill = python_fill
        
        ws.cell(row=idx, column=19, value=py_row['ES_EMA_15'] if pd.notna(py_row['ES_EMA_15']) else "")
        ws.cell(row=idx, column=19).fill = python_fill
        
        ws.cell(row=idx, column=20, value=int(py_row['entry_signal_final']) if pd.notna(py_row['entry_signal_final']) else 0)
        ws.cell(row=idx, column=20).fill = python_fill
        
        # Signal Match (Excel final vs Python final)
        ws.cell(row=idx, column=21, value=f'=P{idx}=T{idx}')
        ws.cell(row=idx, column=21).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # Add Excel EMA formulas after all data is written  
    print("Adding Excel EMA formulas...")
    total_rows = idx  # idx from the last loop iteration
    
    # Calculate EMA using Excel formulas
    for row_num in range(2, total_rows + 1):  # Start from row 2 (first data row)
        current_close = f"B{row_num}"
        
        if row_num == 2:  # First row - EMA = close price
            ws.cell(row=row_num, column=7, value=f'={current_close}')  # EMA_9
            ws.cell(row=row_num, column=8, value=f'={current_close}')  # EMA_15
        else:
            # EMA = (Close * Alpha) + (Previous EMA * (1 - Alpha))
            # Alpha_9 = 2/(9+1) = 0.2, Alpha_15 = 2/(15+1) = 0.125
            prev_ema9 = f"G{row_num-1}"
            prev_ema15 = f"H{row_num-1}"
            
            ema9_formula = f'=IF({current_close}="", {prev_ema9}, ({current_close}*0.2)+({prev_ema9}*0.8))'
            ema15_formula = f'=IF({current_close}="", {prev_ema15}, ({current_close}*0.125)+({prev_ema15}*0.875))'
            
            ws.cell(row=row_num, column=7, value=ema9_formula)
            ws.cell(row=row_num, column=8, value=ema15_formula)
    
    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 15)
        ws.column_dimensions[column].width = adjusted_width
    
    # Add RSI Calculation Debug sheet
    rsi_sheet = wb.create_sheet("RSI Debug")
    
    # RSI Debug headers
    rsi_headers = [
        "Date", "ES_Close", "Price_Change", "Gain", "Loss", 
        "Avg_Gain_2", "Avg_Loss_2", "RS", "RSI_Excel", "RSI_Python", "Difference"
    ]
    
    # Write RSI debug headers
    for col, header in enumerate(rsi_headers, 1):
        cell = rsi_sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Write RSI debug data
    rsi_data_rows = []
    for idx, (raw_idx, raw_row) in enumerate(raw_data.iterrows(), 2):
        py_row = python_data[python_data['date'] == raw_row['date']]
        if py_row.empty:
            continue
        py_row = py_row.iloc[0]
        
        rsi_sheet.cell(row=idx, column=1, value=raw_row['date'])  # Date
        rsi_sheet.cell(row=idx, column=2, value=raw_row['ES_close'] if pd.notna(raw_row['ES_close']) else "")  # ES_Close
        
        if idx == 2:  # First data row, no previous price
            rsi_sheet.cell(row=idx, column=3, value="")  # Price_Change
            rsi_sheet.cell(row=idx, column=4, value="")  # Gain
            rsi_sheet.cell(row=idx, column=5, value="")  # Loss
            rsi_sheet.cell(row=idx, column=6, value="")  # Avg_Gain_2
            rsi_sheet.cell(row=idx, column=7, value="")  # Avg_Loss_2
            rsi_sheet.cell(row=idx, column=8, value="")  # RS
            rsi_sheet.cell(row=idx, column=9, value="")  # RSI_Excel
        else:
            # Price Change = Current - Previous
            rsi_sheet.cell(row=idx, column=3, value=f'=B{idx}-B{idx-1}')
            
            # Gain = IF(Price_Change > 0, Price_Change, 0)
            rsi_sheet.cell(row=idx, column=4, value=f'=IF(C{idx}>0, C{idx}, 0)')
            
            # Loss = IF(Price_Change < 0, -Price_Change, 0)
            rsi_sheet.cell(row=idx, column=5, value=f'=IF(C{idx}<0, -C{idx}, 0)')
            
            if idx == 3:  # Second data row, only 1 period of data
                rsi_sheet.cell(row=idx, column=6, value=f'=D{idx}')  # Avg_Gain_2 = just the gain
                rsi_sheet.cell(row=idx, column=7, value=f'=E{idx}')  # Avg_Loss_2 = just the loss
            else:  # Third row and beyond, can calculate 2-period average
                rsi_sheet.cell(row=idx, column=6, value=f'=AVERAGE(D{idx-1}:D{idx})')  # Avg_Gain_2
                rsi_sheet.cell(row=idx, column=7, value=f'=AVERAGE(E{idx-1}:E{idx})')  # Avg_Loss_2
            
            # RS = Avg_Gain / Avg_Loss
            rsi_sheet.cell(row=idx, column=8, value=f'=IF(G{idx}=0, 999, F{idx}/G{idx})')
            
            # RSI = 100 - (100 / (1 + RS))
            rsi_sheet.cell(row=idx, column=9, value=f'=100-(100/(1+H{idx}))')
        
        # Python RSI value
        rsi_sheet.cell(row=idx, column=10, value=py_row['ES_RSI_2'] if pd.notna(py_row['ES_RSI_2']) else "")
        
        # Difference
        if idx > 2:
            rsi_sheet.cell(row=idx, column=11, value=f'=IF(AND(I{idx}<>"", J{idx}<>""), I{idx}-J{idx}, "")')
    
    # Auto-fit RSI debug columns
    for col in rsi_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 15)
        rsi_sheet.column_dimensions[column].width = adjusted_width
    
    # Add summary sheet
    summary_sheet = wb.create_sheet("Analysis Summary", 0)
    
    summary_data = [
        ["EXCEL vs PYTHON COMPARISON", ""],
        ["", ""],
        ["", ""],
        ["Correct Signal Logic Flow:", ""],
        ["Step 1:", "Base Rule 1: CNN_FEAR_GREED < 30 -> 1/0"],
        ["Step 2:", "Base Rule 2: ES_RSI_2 > 50 -> 1/0"],
        ["Step 3:", "Cond Rule 1: VIX_close > 20 -> 1/0"],
        ["Step 4:", "Cond Rule 2: ES_RSI_2 > 60 -> 1/0"],
        ["Step 5:", "Cond Rule 3: TRIN_close < 0.9 -> 1/0"],
        ["Step 6:", "Cond Sum: Sum of 3 conditional rules (0-3)"],
        ["Step 7:", "Group Signal: IF(CondSum>=1, 1, 0)"],
        ["Step 8:", "Final Signal: Base1=1 AND Base2=1 AND GroupSignal=1"],
        ["", ""],
        ["Column Guide:", ""],
        ["A-H:", "Raw data + Excel indicators (Date, ES, VIX, TRIN, CNN, Excel_RSI, Excel_EMA_9, Excel_EMA_15)"],
        ["I-P:", "Excel signal calculations (Blue=base rules, Yellow=conditional rules)"],
        ["Q-T:", "Python indicator values (Green background: RSI, EMA_9, EMA_15, Final_Signal)"],
        ["U:", "Signal match comparison (Red background)"],
        ["", ""],
        ["Test Case - Feb 16:", ""],
        ["Base1 (CNN<30):", "29.0 < 30 = 1"],
        ["Base2 (RSI>50):", "91.45 > 50 = 1"],
        ["Cond1 (VIX>20):", "24.29 > 20 = 1"],
        ["Cond2 (RSI>60):", "91.45 > 60 = 1"],
        ["Cond3 (TRIN<0.9):", "1.44 < 0.9 = 0"],
        ["Cond Sum:", "1+1+0 = 2"],
        ["Group Signal:", "IF(2>=1, 1, 0) = 1"],
        ["Final Signal:", "1 AND 1 AND 1 = 1"]
    ]
    
    for row_idx, (item, desc) in enumerate(summary_data, 1):
        summary_sheet.cell(row=row_idx, column=1, value=item)
        summary_sheet.cell(row=row_idx, column=2, value=desc)
        if "CORRECTED" in item:
            summary_sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
    
    summary_sheet.column_dimensions['A'].width = 25
    summary_sheet.column_dimensions['B'].width = 40
    
    # Save workbook
    output_file = "Excel_vs_Python_Comparison.xlsx"
    wb.save(output_file)
    print(f"comparison workbook saved: {output_file}")
    
    return output_file

if __name__ == "__main__":
    create_fixed_comparison_workbook()