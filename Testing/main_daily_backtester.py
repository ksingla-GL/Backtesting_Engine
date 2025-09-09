"""
Main Daily Backtester with Excel Output
Orchestrates the entire backtesting process and generates comprehensive Excel reports
"""
import pandas as pd
import numpy as np
import json
import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
import argparse

from daily_data_handler import DailyDataHandler
from daily_signal_generator import DailySignalGenerator
from daily_backtest_engine import DailyBacktestEngine

class MainDailyBacktester:
    def __init__(self):
        self.logger = self._setup_logger()
        self.data_handler = DailyDataHandler()
        self.signal_generator = DailySignalGenerator()
        self.backtest_engine = DailyBacktestEngine()
        
    def _setup_logger(self):
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('daily_backtest.log'),
                logging.StreamHandler()
            ]
        )
        return logging.getLogger(__name__)
    
    def load_strategy_config(self, strategy_name: str) -> dict:
        """Load strategy configuration from JSON file"""
        config_file = f"../Strategies/{strategy_name}.json"
        
        if not os.path.exists(config_file):
            self.logger.error(f"Strategy config file not found: {config_file}")
            return None
        
        try:
            with open(config_file, 'r') as f:
                config = json.load(f)
            
            self.logger.info(f"Loaded strategy config for: {config.get('strategy_name', strategy_name)}")
            return config
            
        except Exception as e:
            self.logger.error(f"Error loading strategy config: {e}")
            return None
    
    def run_strategy_backtest(self, strategy_name: str, data_file: str = 'top_v2_daily_data_2022.csv') -> dict:
        """Run complete backtest for a strategy"""
        self.logger.info(f"Starting backtest for strategy: {strategy_name}")
        
        # Load strategy configuration
        strategy_config = self.load_strategy_config(strategy_name)
        if not strategy_config:
            return None
        
        # Load data
        data = self.data_handler.load_data(data_file)
        if data is None:
            self.logger.error("Failed to load data")
            return None
        
        # Validate required columns
        required_columns = ['ES_close', 'VIX_close', 'TRIN_close', 'CNN_FEAR_GREED']
        if not self.data_handler.validate_data(required_columns):
            return None
        
        # Forward fill missing data
        self.data_handler.forward_fill_missing_data()
        
        # Calculate indicators
        self.logger.info("Calculating indicators...")
        data_with_indicators = self.signal_generator.calculate_indicators(data, strategy_config)
        
        # Generate signals
        self.logger.info("Generating signals...")
        data_with_signals = self.signal_generator.generate_signals(data_with_indicators, strategy_config)
        
        # Run backtest
        self.logger.info("Running backtest...")
        backtest_results = self.backtest_engine.run_backtest(data_with_signals, strategy_config)
        
        # Compile results
        results = {
            'strategy_name': strategy_name,
            'strategy_config': strategy_config,
            'data_with_signals': data_with_signals,
            'backtest_results': backtest_results,
            'signal_summary': self.signal_generator.get_signal_summary(data_with_signals)
        }
        
        self.logger.info("Backtest completed successfully")
        return results
    
    def create_excel_report(self, results: dict, output_file: str = None) -> str:
        """Create comprehensive Excel report"""
        if not results:
            self.logger.error("No results to export")
            return None
        
        strategy_name = results['strategy_name']
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"Backtest_Report_{strategy_name}_{timestamp}.xlsx"
        
        self.logger.info(f"Creating Excel report: {output_file}")
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        data_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        positive_fill = PatternFill(start_color="D4F6D4", end_color="D4F6D4", fill_type="solid")
        negative_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. Summary Sheet
        self._create_summary_sheet(wb, results, header_font, header_fill, thin_border)
        
        # 2. Daily Data Sheet
        self._create_daily_data_sheet(wb, results, header_font, header_fill, data_fill, thin_border)
        
        # 3. Trade Log Sheet
        self._create_trade_log_sheet(wb, results, header_font, header_fill, positive_fill, negative_fill, thin_border)
        
        # 4. Daily Positions Sheet
        self._create_daily_positions_sheet(wb, results, header_font, header_fill, data_fill, thin_border)
        
        # 5. Signal Analysis Sheet
        self._create_signal_analysis_sheet(wb, results, header_font, header_fill, data_fill, thin_border)
        
        # Save workbook
        wb.save(output_file)
        self.logger.info(f"Excel report saved: {output_file}")
        return output_file
    
    def _create_summary_sheet(self, wb, results, header_font, header_fill, border):
        """Create summary sheet with performance metrics"""
        ws = wb.create_sheet("Summary", 0)
        
        strategy_name = results['strategy_name']
        metrics = results['backtest_results']['performance_metrics']
        initial_capital = results['backtest_results']['initial_capital']
        final_value = results['backtest_results']['final_value']
        
        # Title
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = f"BACKTEST SUMMARY - {strategy_name.upper()}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Strategy info
        row = 3
        ws[f'A{row}'] = "Strategy Name:"
        ws[f'B{row}'] = strategy_name
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        ws[f'A{row}'] = "Position Type:"
        ws[f'B{row}'] = results['strategy_config'].get('position_type', 'N/A')
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        ws[f'A{row}'] = "Data Period:"
        data = results['data_with_signals']
        date_range = f"{data['date'].min().strftime('%Y-%m-%d')} to {data['date'].max().strftime('%Y-%m-%d')}"
        ws[f'B{row}'] = date_range
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        
        # Performance metrics
        row += 2
        ws[f'A{row}'] = "PERFORMANCE METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        
        metrics_data = [
            ("Initial Capital", f"${initial_capital:,.2f}"),
            ("Final Value", f"${final_value:,.2f}"),
            ("Total Return", f"{metrics.get('total_return_pct', 0):.2f}%"),
            ("Total Trades", f"{metrics.get('total_trades', 0)}"),
            ("Winning Trades", f"{metrics.get('winning_trades', 0)}"),
            ("Losing Trades", f"{metrics.get('losing_trades', 0)}"),
            ("Win Rate", f"{metrics.get('win_rate_pct', 0):.1f}%"),
            ("Average Win", f"${metrics.get('avg_win', 0):.2f}"),
            ("Average Loss", f"${metrics.get('avg_loss', 0):.2f}"),
            ("Profit Factor", f"{metrics.get('profit_factor', 0):.2f}"),
            ("Max Drawdown", f"{metrics.get('max_drawdown_pct', 0):.2f}%"),
            ("Average Days Held", f"{metrics.get('avg_days_held', 0):.1f}"),
            ("Best Trade", f"${metrics.get('best_trade', 0):.2f}"),
            ("Worst Trade", f"${metrics.get('worst_trade', 0):.2f}")
        ]
        
        for metric, value in metrics_data:
            row += 1
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
    
    def _create_daily_data_sheet(self, wb, results, header_font, header_fill, data_fill, border):
        """Create daily data sheet with all indicators"""
        ws = wb.create_sheet("Daily Data")
        
        data = results['data_with_signals']
        
        # Write data to sheet
        for r in dataframe_to_rows(data, index=False, header=True):
            ws.append(r)
        
        # Style headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Style data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                if cell.row % 2 == 0:
                    cell.fill = data_fill
        
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 15)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_trade_log_sheet(self, wb, results, header_font, header_fill, positive_fill, negative_fill, border):
        """Create trade log sheet"""
        ws = wb.create_sheet("Trade Log")
        
        trades = results['backtest_results']['trades']
        if not trades:
            ws['A1'] = "No trades generated"
            return
        
        trades_df = pd.DataFrame(trades)
        
        # Remove complex columns that can't be serialized to Excel
        columns_to_remove = ['entry_indicators']
        for col in columns_to_remove:
            if col in trades_df.columns:
                trades_df = trades_df.drop(columns=[col])
        
        # Write data to sheet
        for r in dataframe_to_rows(trades_df, index=False, header=True):
            ws.append(r)
        
        # Style headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Style data rows with conditional formatting for PnL
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                
                # Color code PnL columns
                if 'pnl' in str(ws.cell(1, cell.column).value).lower() and cell.value is not None:
                    try:
                        pnl_value = float(cell.value)
                        if pnl_value > 0:
                            cell.fill = positive_fill
                        elif pnl_value < 0:
                            cell.fill = negative_fill
                    except:
                        pass
        
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_daily_positions_sheet(self, wb, results, header_font, header_fill, data_fill, border):
        """Create daily positions sheet"""
        ws = wb.create_sheet("Daily Positions")
        
        positions = results['backtest_results']['daily_positions']
        if not positions:
            ws['A1'] = "No position data available"
            return
        
        positions_df = pd.DataFrame(positions)
        
        # Write data to sheet
        for r in dataframe_to_rows(positions_df, index=False, header=True):
            ws.append(r)
        
        # Style headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Style data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                if cell.row % 2 == 0:
                    cell.fill = data_fill
        
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 15)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_signal_analysis_sheet(self, wb, results, header_font, header_fill, data_fill, border):
        """Create signal analysis sheet"""
        ws = wb.create_sheet("Signal Analysis")
        
        signal_summary = results['signal_summary']
        
        # Title
        ws['A1'] = "SIGNAL ANALYSIS"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Headers
        headers = ["Signal/Rule", "True Count", "False Count", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Data
        row = 4
        for signal_name, info in signal_summary.items():
            ws.cell(row=row, column=1, value=signal_name).border = border
            ws.cell(row=row, column=2, value=info['true_count']).border = border
            ws.cell(row=row, column=3, value=info['false_count']).border = border
            ws.cell(row=row, column=4, value=f"{info['percentage']:.1f}%").border = border
            
            # Alternate row coloring
            if row % 2 == 0:
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = data_fill
            
            row += 1
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12

def main():
    """Main function for command line usage"""
    parser = argparse.ArgumentParser(description='Run daily backtest with Excel output')
    parser.add_argument('strategy', help='Strategy name (without .json extension)')
    parser.add_argument('--data', default='top_v2_daily_data_2022.csv', help='Data file name')
    parser.add_argument('--output', help='Output Excel file name (auto-generated if not provided)')
    
    args = parser.parse_args()
    
    # Create backtester
    backtester = MainDailyBacktester()
    
    # Run backtest
    results = backtester.run_strategy_backtest(args.strategy, args.data)
    
    if results:
        # Print performance summary
        backtester.backtest_engine.print_performance_summary()
        
        # Create Excel report
        excel_file = backtester.create_excel_report(results, args.output)
        print(f"\nExcel report created: {excel_file}")
        
        # Save trade log CSV
        if results['backtest_results']['trades']:
            trades_df = pd.DataFrame(results['backtest_results']['trades'])
            csv_file = f"trade_log_{args.strategy}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            trades_df.to_csv(csv_file, index=False)
            print(f"Trade log CSV saved: {csv_file}")
    
    else:
        print("Backtest failed")

if __name__ == "__main__":
    main()